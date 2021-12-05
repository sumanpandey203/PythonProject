from openpyxl import Workbook
from openpyxl import load_workbook



##############################


myOddList = []
myEvenList = []
myTextList = []
myEven_even = []
myEven_odd= []
myEven_odd_en = []

#############################

#open excel
path = r"C:\Users\suman.pandey\Desktop\My_work\Pyth\MindMap.xlsx"
wb_object = load_workbook(path)

sheet1_obj = wb_object.active


cell_obj = sheet1_obj.cell(row = 1, column = 1)

print(cell_obj.value)

for col in range(1,42):

    for r in range(1,4):
        cell_obj = sheet1_obj.cell(row = r, column = col)
        myOddList.append(cell_obj.value)
print(myOddList)

wb = Workbook()
sheet = wb.create_sheet(title='MindMap')

for idx in range(123):
    cell = sheet.cell(row = idx+1, column = 1)
    cell.value= myOddList[idx]
wb.save(r"C:\Users\suman.pandey\Desktop\My_work\Pyth\MindMap1.xlsx")

# wb.save(r"C:\Users\suman.pandey\Desktop\My_work\Pyth\MindMap.xlsx")
# sheet = wb.active
# sheet.title = 'mindmap'
# sheet1 = wb.create_sheet(title='Tesiko_Info')
# sheet2 = wb.create_sheet(title='Tesiko_Katalyst')